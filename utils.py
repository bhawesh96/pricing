import openpyxl
import datetime
import os
import json

def all_products(source, name, qty, mrp, sp):
    file_name = '{}_{}.json'.format(source, datetime.datetime.today().date().isoformat())
    all_prods = {'source': source, 'iso_datetime': datetime.datetime.today().isoformat(), 'data': []}

    for a, b, c, d in zip(name, qty, mrp, sp):
        all_prods['data'].append({
            'name': a,
            'qty': b,
            'mrp': c,
            'sp': d
        })

    with open(os.path.join('all_jsons', file_name), 'w') as f:
        json.dump(all_prods, f)

    return file_name

def write_lists_to_excel(source, name, qty, mrp, sp):
    filename = "{}.xlsx".format(datetime.datetime.today().date().isoformat())
    # Load the workbook
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        # Create a new workbook if the file doesn't exist
        workbook = openpyxl.Workbook()
        # Remove the default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    # Create a new sheet
    sheet = workbook.create_sheet(source)

    # Write the header names
    header_list = ['Item', 'Quantity', 'MRP', 'SP']
    for col_num, header in enumerate(header_list, start=1):
        sheet.cell(row=1, column=col_num).value = header

    # Write the lists to the columns
    for index, item in enumerate(name, start=2):
        sheet.cell(row=index, column=1).value = item

    for index, item in enumerate(qty, start=2):
        sheet.cell(row=index, column=2).value = item

    for index, item in enumerate(mrp, start=2):
        sheet.cell(row=index, column=3).value = item

    for index, item in enumerate(sp, start=2):
        sheet.cell(row=index, column=4).value = item

    # Save the workbook
    workbook.save(filename)
    print(f"Data written successfully to '{source}' sheet in '{filename}'.")

    file_name = all_products(source, name, qty, mrp, sp)
    print(f"Data written successfully to '{file_name}'.")